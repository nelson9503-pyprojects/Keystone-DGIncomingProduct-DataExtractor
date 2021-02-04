import tkinter as tk
from tkinter import ttk
from tkinter import simpledialog
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
import threading
import json
import os


class Extractor:

    def __init__(self, fileconfig: dict, colsconfig: dict):
        self.fileconfig = fileconfig
        for item in colsconfig:
            colsconfig[item] = openpyxl.utils.column_index_from_string(
                colsconfig[item])
        self.colsconfig = colsconfig
        self.open_map()
        self.extract_excel()
        self.mapping_results()
        self.export_excel()

    def extract_excel(self):

        path = self.fileconfig["path"]
        sheet = self.fileconfig["sheet"]
        startrow = int(self.fileconfig["startrow"])

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sh = wb[sheet]
        table = []
        for row in sh.rows:
            cells = []
            for cell in row:
                cells.append(cell.value)
            table.append(cells)

        self.result = {}
        order = 0

        row = startrow - 1
        while not table[row][self.colsconfig["spec"]-1] in [None, ""]:
            self.result[order] = {}
            for item in self.colsconfig.keys():
                self.result[order][item] = table[row][self.colsconfig[item]-1]
            order += 1
            row += 1

        wb.close()

    def open_map(self):
        path = "./keystone_dg_mapping.json"
        if not os.path.exists(path):
            with open(path, 'w') as f:
                f.write(json.dumps({}))
        with open(path, 'r') as f:
            self.map = json.loads(f.read())

    def save_map(self):
        path = "./keystone_dg_mapping.json"
        with open(path, 'w') as f:
            f.write(json.dumps(self.map))

    def mapping_results(self):

        for order in self.result:
            color = self.result[order]["color"]
            if color == None:
                color = ""
            dgcode = self.result[order]["spec"] + "[" + color + "]"
            if not dgcode in self.map:
                hkcode = simpledialog.askstring(
                    "Please provide hk code",
                    "Please provide HK code for:\n\n{}\n{}\n\n\t1. Press cancel to skip.\n\t2. Enter 'end' to terminate mapping".format(
                        dgcode, self.result[order]["packing"])
                )
                if not hkcode in [None, "", "end"]:
                    self.map[dgcode] = hkcode
                    self.save_map()
            else:
                hkcode = self.map[dgcode]
            if hkcode == "end":
                break
            elif not hkcode in [None, "", "end"]:
                self.result[order]["M18"] = hkcode

    def export_excel(self):
        path = filedialog.asksaveasfilename(
            title="Save extracting output",
            filetypes=(("Excel Files", "*.xlsx"),),
            defaultextension=".xlsx"
        )
        if not path in [None, ""]:
            wb = openpyxl.Workbook()
            sh = wb.create_sheet("Results", 0)
            col = 1
            for header in self.colsconfig:
                sh.cell(1, col).value = header
                col += 1
            sh.cell(1, col).value = "M18"
            row = 2
            for order in self.result:
                result = self.result[order]
                col = 1
                for item in self.colsconfig.keys():
                    sh.cell(row, col).value = result[item]
                    col += 1
                if "M18" in result:
                    sh.cell(row, col).value = result["M18"]
                row += 1
            wb.save(path)
            messagebox.showinfo("Export Excel", "Export Success!")
        else:
            messagebox.showerror("Export Excel", "Export Failed!")
